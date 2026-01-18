from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from config import get_config
from models import db, User, Group, Student, Attendance
from auth import auth_bp, init_auth
from utils import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

app = Flask(__name__)
app.config.from_object(get_config())

# Database ni sozlash
db.init_app(app)

# Authentication ni sozlash
init_auth(app)
app.register_blueprint(auth_bp)

# Database yaratish
with app.app_context():
    db.create_all()
    
    # Admin yaratish (agar yo'q bo'lsa)
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print("Admin yaratildi: username=admin, password=admin123")

# ==================== ROUTES ====================

@app.route('/')
def index():
    """Bosh sahifa - dashboard ga yo'naltirish"""
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('auth.login'))

# Main blueprint
from flask import Blueprint
main_bp = Blueprint('main', __name__)

@main_bp.route('/dashboard')
@login_required
def dashboard():
    """Dashboard - real-time statistika"""
    stats = get_dashboard_stats()
    return render_template('dashboard.html', stats=stats)

@main_bp.route('/students/add', methods=['GET', 'POST'])
@login_required
def students_add():
    """Talaba qo'shish - bitta oyna"""
    if request.method == 'POST':
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        patronymic = request.form.get('patronymic', '').strip()
        group_name = request.form.get('group_name', '').strip()
        
        if not all([first_name, last_name, patronymic, group_name]):
            flash('Barcha maydonlarni to\'ldiring!', 'danger')
            return redirect(url_for('main.students_add'))
        
        # Guruhni olish yoki yaratish
        group = get_or_create_group(group_name)
        
        # Talaba yaratish
        student = Student(
            first_name=first_name,
            last_name=last_name,
            patronymic=patronymic,
            group_id=group.id
        )
        db.session.add(student)
        db.session.commit()
        
        flash(f'Talaba qo\'shildi: {student.full_name} ({group.name})', 'success')
        return redirect(url_for('main.students_add'))
    
    # Barcha guruhlarni olish (autocomplete uchun)
    groups = Group.query.order_by(Group.name).all()
    return render_template('student_add.html', groups=groups)

@main_bp.route('/students/delete/<int:student_id>', methods=['POST'])
@login_required
def students_delete(student_id):
    """Talabani o'chirish"""
    student = Student.query.get_or_404(student_id)
    student_name = student.full_name
    
    db.session.delete(student)
    db.session.commit()
    
    flash(f'Talaba o\'chirildi: {student_name}', 'info')
    return redirect(request.referrer or url_for('main.dashboard'))

@main_bp.route('/attendance/mark', methods=['GET', 'POST'])
@login_required
def attendance_mark():
    """Davomat belgilash - jadval tizimi"""
    if request.method == 'POST':
        # AJAX so'rov - davomat saqlash
        data = request.get_json()
        student_id = data.get('student_id')
        target_date = datetime.strptime(data.get('date'), '%Y-%m-%d').date()
        hours = data.get('hours', [])  # 7 elementli ro'yxat
        
        # Davomat yozuvini olish yoki yaratish
        attendance = get_or_create_attendance(student_id, target_date)
        
        # Soatlarni saqlash
        attendance.set_hours_list(hours)
        attendance.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Saqlandi'})
    
    # GET so'rov - sahifani ko'rsatish
    groups = Group.query.order_by(Group.name).all()
    current_date = get_current_date()
    
    selected_group_id = request.args.get('group_id', type=int)
    students = []
    attendances = {}
    
    if selected_group_id:
        # Guruhga kirish statistikasini yangilash
        update_group_access(selected_group_id)
        
        # Talabalarni alfabit tartibida olish
        students = get_students_alphabetically(selected_group_id)
        
        # Har bir talaba uchun bugungi davomatni olish
        for student in students:
            att = get_or_create_attendance(student.id, current_date)
            attendances[student.id] = att.get_hours_list()
    
    return render_template(
        'attendance_mark.html',
        groups=groups,
        selected_group_id=selected_group_id,
        students=students,
        attendances=attendances,
        current_date=current_date,
        current_datetime=get_current_datetime()
    )

@main_bp.route('/reports/groups')
@login_required
def reports_groups():
    """Hisobotlar - guruhlar ro'yxati"""
    groups = Group.query.order_by(Group.name).all()
    return render_template('reports_groups.html', groups=groups)

@main_bp.route('/reports/group/<int:group_id>')
@login_required
def reports_group(group_id):
    """Hisobot - guruh ichida"""
    group = Group.query.get_or_404(group_id)
    
    # Guruhga kirish statistikasini yangilash
    update_group_access(group_id)
    
    # Sana tanlanganmi?
    selected_date_str = request.args.get('date')
    if selected_date_str:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    else:
        selected_date = get_current_date()
    
    # Talabalarni olish
    students = get_students_alphabetically(group_id)
    
    # Har bir talaba uchun davomat
    attendances = {}
    for student in students:
        att = Attendance.query.filter_by(
            student_id=student.id,
            date=selected_date
        ).first()
        if att:
            attendances[student.id] = att.get_hours_list()
        else:
            attendances[student.id] = [None] * 7
    
    return render_template(
        'reports_group.html',
        group=group,
        students=students,
        attendances=attendances,
        selected_date=selected_date
    )

@main_bp.route('/reports/student/<int:student_id>')
@login_required
def reports_student(student_id):
    """Hisobot - talabaning to'liq tarixi"""
    student = Student.query.get_or_404(student_id)
    
    # Talabaning barcha davomat yozuvlarini olish
    history = get_student_attendance_history(student_id)
    
    # Jami g'oyiblar
    total_absences = calculate_total_absences(student_id)
    
    return render_template(
        'reports_student.html',
        student=student,
        history=history,
        total_absences=total_absences
    )

@main_bp.route('/reports/student/<int:student_id>/delete-record', methods=['POST'])
@login_required
def delete_absence_record(student_id):
    """G'oyibni o'chirish (✅ ga o'zgartirish)"""
    data = request.get_json()
    attendance_id = data.get('attendance_id')
    hour_num = data.get('hour_num')  # 1-7
    
    attendance = Attendance.query.get_or_404(attendance_id)
    
    # Soatni ✅ ga o'zgartirish
    hours = attendance.get_hours_list()
    hours[hour_num - 1] = True
    attendance.set_hours_list(hours)
    attendance.updated_at = datetime.utcnow()
    db.session.commit()
    
    # Yangilangan g'oyiblar sonini qaytarish
    new_total = calculate_total_absences(student_id)
    
    return jsonify({
        'success': True,
        'new_total': new_total,
        'message': 'O\'zgartirildi'
    })

@main_bp.route('/export/group/<int:group_id>')
@login_required
def export_group(group_id):
    """Guruhni Excel ga export qilish"""
    group = Group.query.get_or_404(group_id)
    
    # Sana oralig'ini olish
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        # Default: oxirgi 30 kun
        end_date = get_current_date()
        start_date = end_date - timedelta(days=30)
    
    # Excel fayl yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = f"{group.name}"
    
    # Sarlavha
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"DAVOMAT HISOBOTI - {group.name}"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:K2')
    date_cell = ws['A2']
    date_cell.value = f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    date_cell.alignment = Alignment(horizontal='center')
    
    # Bo'sh qator
    current_row = 4
    
    # Jadval sarlavhasi
    headers = ['№', 'ISM', 'FAMILIYA', 'OTCHESTVO', '1', '2', '3', '4', '5', '6', '7', 'JAMI G\'OYIB']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 1
    
    # Talabalarni olish
    students = get_students_alphabetically(group_id)
    
    # Har bir sana uchun
    delta = end_date - start_date
    
    for idx, student in enumerate(students, 1):
        # Talaba ma'lumotlari
        ws.cell(row=current_row, column=1).value = idx
        ws.cell(row=current_row, column=2).value = student.first_name
        ws.cell(row=current_row, column=3).value = student.last_name
        ws.cell(row=current_row, column=4).value = student.patronymic
        
        total_absent = 0
        
        # Har bir sana uchun davomat
        for i in range(delta.days + 1):
            check_date = start_date + timedelta(days=i)
            attendance = Attendance.query.filter_by(
                student_id=student.id,
                date=check_date
            ).first()
            
            if attendance:
                total_absent += attendance.count_absent()
        
        # Oxirgi davomat (eng so'nggi kun)
        last_attendance = Attendance.query.filter_by(
            student_id=student.id,
            date=end_date
        ).first()
        
        if last_attendance:
            hours = last_attendance.get_hours_list()
            for col, hour_status in enumerate(hours, 5):
                cell = ws.cell(row=current_row, column=col)
                if hour_status is True:
                    cell.value = '✅'
                elif hour_status is False:
                    cell.value = '❌'
                else:
                    cell.value = '-'
                cell.alignment = Alignment(horizontal='center')
        
        # Jami g'oyib
        ws.cell(row=current_row, column=12).value = total_absent
        ws.cell(row=current_row, column=12).font = Font(bold=True)
        ws.cell(row=current_row, column=12).alignment = Alignment(horizontal='center')
        
        current_row += 1
    
    # Ustunlar kengligini sozlash
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 5
    ws.column_dimensions['L'].width = 15
    
    # Faylni xotiraga saqlash
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Davomat_{group.name}_{start_date}_{end_date}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# Blueprint ni ro'yxatdan o'tkazish
app.register_blueprint(main_bp)

# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    flash('Sahifa topilmadi!', 'danger')
    return redirect(url_for('main.dashboard'))

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    flash('Ichki server xatolik!', 'danger')
    return redirect(url_for('main.dashboard'))

# ==================== CONTEXT PROCESSORS ====================

@app.context_processor
def utility_processor():
    """Barcha template larda foydalanish uchun"""
    return {
        'now': datetime.now(),
        'today': date.today()
    }

# ==================== PRODUCTION SETTINGS ====================

if __name__ == '__main__':
    # Development
    app.run(debug=True, host='0.0.0.0', port=5656)
else:
    # Production (Render, Heroku, etc.)
    import logging
    logging.basicConfig(level=logging.INFO)
