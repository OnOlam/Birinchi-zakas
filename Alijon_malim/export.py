"""
Excel Export Module
Attendance hisobotlarini Excel formatga export qilish
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


class AttendanceExcelExporter:
    """
    Davomat hisobotlarini Excel'ga export qilish
    """
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Ranglar
        self.colors = {
            'header': 'FF4472C4',      # Ko'k
            'present': 'FF70AD47',     # Yashil
            'absent': 'FFFF0000',      # Qizil
            'total': 'FFFFC000',       # Sariq
            'group_header': 'FF5B9BD5' # Och ko'k
        }
        
        # Fontlar
        self.fonts = {
            'header': Font(name='Arial', size=14, bold=True, color='FFFFFF'),
            'subheader': Font(name='Arial', size=12, bold=True, color='FFFFFF'),
            'title': Font(name='Arial', size=16, bold=True),
            'normal': Font(name='Arial', size=11),
            'bold': Font(name='Arial', size=11, bold=True)
        }
    
    def _apply_border(self, cell):
        """
        Cell'ga border qo'shish
        """
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    def _set_column_width(self, column, width):
        """
        Ustun kengligini sozlash
        """
        self.ws.column_dimensions[get_column_letter(column)].width = width
    
    def export_daily_report(self, date, groups_data):
        """
        Kunlik hisobot - barcha guruhlar
        
        Args:
            date: Hisobot sanasi
            groups_data: Guruhlar ma'lumotlari
            
        Returns:
            BytesIO: Excel fayl
        """
        self.ws.title = f"Davomat {date.strftime('%d.%m.%Y')}"
        
        current_row = 1
        
        # Sarlavha
        self.ws.merge_cells(f'A{current_row}:E{current_row}')
        title_cell = self.ws[f'A{current_row}']
        title_cell.value = f"DAVOMAT HISOBOTI - {date.strftime('%d.%m.%Y')}"
        title_cell.font = self.fonts['title']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2
        
        # Umumiy statistika
        self.ws.merge_cells(f'A{current_row}:E{current_row}')
        summary_cell = self.ws[f'A{current_row}']
        
        total_students = sum(g['total'] for g in groups_data)
        total_present = sum(g['present'] for g in groups_data)
        total_absent = sum(g['absent'] for g in groups_data)
        percentage = (total_present / total_students * 100) if total_students > 0 else 0
        
        summary_cell.value = f"Jami: {total_students} talaba | Kelgan: {total_present} | Kelmagan: {total_absent} | Foiz: {percentage:.1f}%"
        summary_cell.font = self.fonts['bold']
        summary_cell.fill = PatternFill(start_color=self.colors['total'], fill_type='solid')
        summary_cell.alignment = Alignment(horizontal='center')
        self._apply_border(summary_cell)
        current_row += 2
        
        # Har bir guruh uchun
        for group_data in groups_data:
            # Guruh nomi
            self.ws.merge_cells(f'A{current_row}:E{current_row}')
            group_cell = self.ws[f'A{current_row}']
            group_cell.value = f"📁 {group_data['group_name']}"
            group_cell.font = self.fonts['subheader']
            group_cell.fill = PatternFill(start_color=self.colors['group_header'], fill_type='solid')
            group_cell.alignment = Alignment(horizontal='left')
            self._apply_border(group_cell)
            current_row += 1
            
            # Header
            headers = ['#', 'Ism', 'Otchestvo', 'Familiya', 'Status', 'Izoh']
            for col, header in enumerate(headers, start=1):
                cell = self.ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = self.fonts['header']
                cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                self._apply_border(cell)
            current_row += 1
            
            # Talabalar
            for idx, student in enumerate(group_data['students'], start=1):
                # Raqam
                cell = self.ws.cell(row=current_row, column=1)
                cell.value = idx
                cell.alignment = Alignment(horizontal='center')
                self._apply_border(cell)
                
                # Ism
                cell = self.ws.cell(row=current_row, column=2)
                cell.value = student['first_name']
                self._apply_border(cell)
                
                # Otchestvo
                cell = self.ws.cell(row=current_row, column=3)
                cell.value = student.get('middle_name', '') or ''
                self._apply_border(cell)
                
                # Familiya
                cell = self.ws.cell(row=current_row, column=4)
                cell.value = student['last_name']
                self._apply_border(cell)
                
                # Status
                cell = self.ws.cell(row=current_row, column=5)
                if student['status'] == 'present':
                    cell.value = "✅ Keldi"
                    cell.fill = PatternFill(start_color=self.colors['present'], fill_type='solid')
                    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                elif student['status'] == 'absent':
                    cell.value = "❌ Kelmadi"
                    cell.fill = PatternFill(start_color=self.colors['absent'], fill_type='solid')
                    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                else:
                    cell.value = "⏳ Belgilanmagan"
                    cell.fill = PatternFill(start_color='FFCCCCCC', fill_type='solid')
                
                cell.alignment = Alignment(horizontal='center')
                self._apply_border(cell)
                
                # Izoh (bo'sh)
                cell = self.ws.cell(row=current_row, column=6)
                cell.value = ""
                self._apply_border(cell)
                
                current_row += 1
            
            # Guruh statistikasi
            self.ws.merge_cells(f'A{current_row}:B{current_row}')
            stat_cell = self.ws.cell(row=current_row, column=1)
            stat_cell.value = f"Jami: {group_data['total']}"
            stat_cell.font = self.fonts['bold']
            stat_cell.fill = PatternFill(start_color='FFE7E6E6', fill_type='solid')
            self._apply_border(stat_cell)
            
            cell = self.ws.cell(row=current_row, column=3)
            cell.value = f"Kelgan: {group_data['present']}"
            cell.font = self.fonts['bold']
            cell.fill = PatternFill(start_color='FFE7E6E6', fill_type='solid')
            self._apply_border(cell)
            
            cell = self.ws.cell(row=current_row, column=4)
            cell.value = f"Kelmagan: {group_data['absent']}"
            cell.font = self.fonts['bold']
            cell.fill = PatternFill(start_color='FFE7E6E6', fill_type='solid')
            self._apply_border(cell)
            
            self.ws.merge_cells(f'E{current_row}:F{current_row}')
            cell = self.ws.cell(row=current_row, column=5)
            group_percentage = (group_data['present'] / group_data['total'] * 100) if group_data['total'] > 0 else 0
            cell.value = f"{group_percentage:.1f}%"
            cell.font = self.fonts['bold']
            cell.fill = PatternFill(start_color='FFE7E6E6', fill_type='solid')
            self._apply_border(cell)
            
            current_row += 2  # Bo'sh qator
        
        # Ustun kengliklari
        self._set_column_width(1, 6)   # #
        self._set_column_width(2, 18)  # Ism
        self._set_column_width(3, 18)  # Otchestvo
        self._set_column_width(4, 18)  # Familiya
        self._set_column_width(5, 18)  # Status
        self._set_column_width(6, 30)  # Izoh
        
        # BytesIO'ga saqlash
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output
    
    def export_group_report(self, date, group_name, students_data):
        """
        Bitta guruh uchun batafsil hisobot
        
        Args:
            date: Sana
            group_name: Guruh nomi
            students_data: Talabalar ma'lumotlari
            
        Returns:
            BytesIO: Excel fayl
        """
        self.ws.title = f"{group_name}"
        
        current_row = 1
        
        # Sarlavha
        self.ws.merge_cells(f'A{current_row}:D{current_row}')
        title_cell = self.ws[f'A{current_row}']
        title_cell.value = f"GURUH: {group_name}"
        title_cell.font = self.fonts['title']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Sana
        self.ws.merge_cells(f'A{current_row}:D{current_row}')
        date_cell = self.ws[f'A{current_row}']
        date_cell.value = f"Sana: {date.strftime('%d.%m.%Y')}"
        date_cell.font = self.fonts['bold']
        date_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Header
        headers = ['#', 'Ism Familiya', 'Status', 'Vaqt']
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            self._apply_border(cell)
        current_row += 1
        
        # Talabalar
        present_count = 0
        absent_count = 0
        
        for idx, student in enumerate(students_data, start=1):
            # Raqam
            cell = self.ws.cell(row=current_row, column=1)
            cell.value = idx
            cell.alignment = Alignment(horizontal='center')
            self._apply_border(cell)
            
            # Ism Familiya
            cell = self.ws.cell(row=current_row, column=2)
            cell.value = f"{student['first_name']} {student['last_name']}"
            self._apply_border(cell)
            
            # Status
            cell = self.ws.cell(row=current_row, column=3)
            if student['status'] == 'present':
                cell.value = "✅ Keldi"
                cell.fill = PatternFill(start_color=self.colors['present'], fill_type='solid')
                cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                present_count += 1
            elif student['status'] == 'absent':
                cell.value = "❌ Kelmadi"
                cell.fill = PatternFill(start_color=self.colors['absent'], fill_type='solid')
                cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                absent_count += 1
            else:
                cell.value = "⏳ Belgilanmagan"
            
            cell.alignment = Alignment(horizontal='center')
            self._apply_border(cell)
            
            # Vaqt (bo'sh)
            cell = self.ws.cell(row=current_row, column=4)
            cell.value = ""
            self._apply_border(cell)
            
            current_row += 1
        
        # Statistika
        current_row += 1
        total = len(students_data)
        percentage = (present_count / total * 100) if total > 0 else 0
        
        self.ws.merge_cells(f'A{current_row}:D{current_row}')
        stat_cell = self.ws[f'A{current_row}']
        stat_cell.value = f"Jami: {total} | Kelgan: {present_count} | Kelmagan: {absent_count} | Foiz: {percentage:.1f}%"
        stat_cell.font = self.fonts['bold']
        stat_cell.fill = PatternFill(start_color=self.colors['total'], fill_type='solid')
        stat_cell.alignment = Alignment(horizontal='center')
        self._apply_border(stat_cell)
        
        # Ustun kengliklari
        self._set_column_width(1, 6)   # #
        self._set_column_width(2, 30)  # Ism Familiya
        self._set_column_width(3, 18)  # Status
        self._set_column_width(4, 15)  # Vaqt
        
        # BytesIO'ga saqlash
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output


# Helper function
def generate_filename(prefix, date, group_name=None):
    """
    Excel fayl nomi generatsiya qilish
    
    Args:
        prefix: Fayl prefiksi
        date: Sana
        group_name: Guruh nomi (ixtiyoriy)
        
    Returns:
        str: Fayl nomi
    """
    date_str = date.strftime('%Y-%m-%d')
    
    if group_name:
        # Guruh nomidagi maxsus belgilarni tozalash
        safe_group_name = "".join(c for c in group_name if c.isalnum() or c in (' ', '-', '_')).strip()
        return f"{prefix}_{safe_group_name}_{date_str}.xlsx"
    else:
        return f"{prefix}_{date_str}.xlsx"


if __name__ == '__main__':
    # Test
    print("=== EXCEL EXPORT MODULE TEST ===")
    print("openpyxl installed: OK")
